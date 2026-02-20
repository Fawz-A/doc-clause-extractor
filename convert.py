import os
import re
from pathlib import Path
from typing import List, Literal, Optional

import pdfplumber
from PIL import Image
import pandas as pd
import pytesseract

from sqlalchemy import create_engine, text

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

OutputTarget = Literal["excel", "postgres"]


# -----------------------------
# TEXT EXTRACTION
# -----------------------------

def extract_text_from_pdf(pdf_path: Path) -> List[str]:
    pages_text: List[str] = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text_content = page.extract_text() or ""
            pages_text.append(text_content.strip())

    return pages_text


def extract_text_from_image(image_path: Path) -> List[str]:
    img = Image.open(image_path)
    img = img.convert("L")

    text_content = pytesseract.image_to_string(
        img,
        config="--psm 6"
    )

    return [text_content.strip()]


# -----------------------------
# CLAUSE-BASED PROCESSING
# -----------------------------

def process_text_pages(text_pages: List[str]) -> pd.DataFrame:
    rows = []

    clause_pattern = re.compile(r"^\s*(\d+(?:\.\d+)*)\s+(.*)")

    current_clause = None
    current_text = ""

    for page_content in text_pages:
        lines = page_content.splitlines()

        for line in lines:
            line = line.strip()
            if not line:
                continue

            match = clause_pattern.match(line)

            if match:
                if current_clause:
                    rows.append({
                        "clause_number": current_clause,
                        "content": current_text.strip(),
                    })

                current_clause = match.group(1)
                current_text = match.group(2)

            else:
                if current_clause:
                    current_text += " " + line

    if current_clause:
        rows.append({
            "clause_number": current_clause,
            "content": current_text.strip(),
        })

    df = pd.DataFrame(rows)

    # ✅ Add description column (empty for manual fill)
    df["description"] = ""

    return df


# -----------------------------
# SAVE TO EXCEL (Formatted)
# -----------------------------

def save_to_excel(text_pages: List[str], output_path: Path) -> None:
    df = process_text_pages(text_pages)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Write raw Excel
    df.to_excel(output_path, index=False)

    # Load workbook for formatting
    wb = load_workbook(output_path)
    ws = wb.active

    # ✅ Freeze header row
    ws.freeze_panes = "A2"

    # ✅ Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # ✅ Enable wrap text for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ✅ Auto-adjust column widths (with limit)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    wb.save(output_path)

    print(f"Saved formatted Excel file to: {output_path}")


# -----------------------------
# SAVE TO POSTGRES
# -----------------------------

def save_to_postgres(
    text_pages: List[str],
    connection_url: str,
    table_name: str = "document_content",
) -> None:

    df = process_text_pages(text_pages)
    engine = create_engine(connection_url)

    create_table_sql = f"""
    CREATE TABLE IF NOT EXISTS {table_name} (
        id SERIAL PRIMARY KEY,
        clause_number TEXT,
        content TEXT,
        description TEXT
    );
    """

    with engine.begin() as conn:
        conn.execute(text(create_table_sql))

        insert_sql = text(
            f"""
            INSERT INTO {table_name}
            (clause_number, content, description)
            VALUES (:clause, :content, :description)
            """
        )

        for _, row in df.iterrows():
            conn.execute(
                insert_sql,
                {
                    "clause": row["clause_number"],
                    "content": row["content"],
                    "description": row["description"],
                },
            )

    print(f"Inserted {len(df)} clauses into table '{table_name}'.")


# -----------------------------
# MAIN CONVERTER
# -----------------------------

def convert_file(
    input_path: str,
    target: OutputTarget,
    excel_output: Optional[str] = None,
    postgres_url: Optional[str] = None,
    postgres_table: str = "document_content",
) -> None:

    path = Path(input_path)

    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    ext = path.suffix.lower()

    if ext == ".pdf":
        text_pages = extract_text_from_pdf(path)
    elif ext in {".jpg", ".jpeg", ".png"}:
        text_pages = extract_text_from_image(path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    if target == "excel":
        if not excel_output:
            excel_output = str(path.with_suffix(".xlsx"))
        save_to_excel(text_pages, Path(excel_output))

    elif target == "postgres":
        if not postgres_url:
            raise ValueError("postgres_url is required when target='postgres'")
        save_to_postgres(text_pages, postgres_url, table_name=postgres_table)

    else:
        raise ValueError(f"Unknown target: {target}")


# -----------------------------
# ENTRY POINT
# -----------------------------

if __name__ == "__main__":
    print("Starting conversion...")

    convert_file(
        "input.pdf",
        target="excel",
        excel_output="technical_specs_output.xlsx",
    )

    print("Done.")
    